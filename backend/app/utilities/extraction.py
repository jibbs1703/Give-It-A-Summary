"""Give It A Summary - Extraction Utilities Module."""

import csv
from pathlib import Path

import pdfplumber
from docx import Document
from openpyxl import load_workbook

EXTRACTORS = {}


def register_extractor(extension):
    """
    Decorator to register an extractor function for a given file extension.

    Args:
    ----------
    extension : str
        File extension (e.g., '.pdf', '.txt', '.docx').
        Must be lowercase and include the leading dot.

    Returns
    -------
    callable
        Decorator that registers the function in the EXTRACTORS registry.
    """
    def decorator(func):
        EXTRACTORS[extension] = func
        return func
    return decorator


@register_extractor(".pdf")
def extract_text_from_pdf(pdf_path, pages=None):
    """
    Extract text from a PDF file using pdfplumber.
    """
    text = []
    with pdfplumber.open(Path(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            if pages is None or i in pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
    return "\n".join(text)


@register_extractor(".txt")
def extract_text_from_txt(txt_path):
    """
    Extract text from a plain text file with encoding fallback.
    """
    try:
        with open(txt_path, encoding="utf-8") as file:
            return file.read()
    except UnicodeDecodeError:
        with open(txt_path, encoding="latin-1") as file:
            return file.read()


@register_extractor(".xlsx")
@register_extractor(".xls")
def extract_text_from_excel(excel_path):
    """
    Extract text from Excel workbooks (.xls/.xlsx).
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text.append(f"Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                text.append(row_text)
        text.append("")
    return "\n".join(text)


@register_extractor(".docx")
def extract_text_from_docx(docx_path):
    """
    Extract text from Microsoft Word documents (.docx).
    """
    doc = Document(docx_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


@register_extractor(".csv")
def extract_text_from_csv(csv_path, delimiter=","):
    """
    Extract text from CSV files.
    """
    text = []
    with open(csv_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile, delimiter=delimiter)
        for row in reader:
            row_text = " ".join(field for field in row if field.strip())
            if row_text:
                text.append(row_text)
    return "\n".join(text)


def extract_text(file_path, **kwargs):
    """
    Unified text extraction dispatcher using Registry Pattern.

    Parameters
    ----------
    file_path : str or Path
        Path to the input file.
    **kwargs : dict
        Optional keyword arguments passed to the registered extractor
        (e.g., `pages` for PDF, `delimiter` for CSV).

    Returns
    -------
    str
        Extracted textual content.

    Raises
    ------
    ValueError
        If no extractor is registered for the file extension.

    Design Notes
    ------------
    - Registry Pattern decouples dispatch logic from extractor implementations.
    - Adding support for new formats requires only:
        1. Writing a new extractor function.
        2. Decorating it with `@register_extractor(".ext")`.
    - This design scales well in enterprise pipelines where file types evolve.
    """
    ext = Path(file_path).suffix.lower()
    extractor = EXTRACTORS.get(ext)
    if not extractor:
        raise ValueError(f"No extractor registered for extension: {ext}")
    return extractor(file_path, **kwargs)
