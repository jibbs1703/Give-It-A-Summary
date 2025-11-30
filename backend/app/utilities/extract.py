"""Give It A Summary - Extraction Utilities Module."""

import csv
from pathlib import Path

import pdfplumber
from docx import Document
from langgraph.tools import tool
from openpyxl import load_workbook

from app.models.agents import ExtractTextInputs
from app.utilities.logs import get_logger

logger = get_logger(__name__)

EXTRACTORS = {}


def register_extractor(extension) -> callable:
    """
    Decorator to register extractor functions for specific file extensions.

    Args:
        extension (str): File extension (e.g., ".pdf", ".txt").

    Returns:
        callable: Decorator function.
    """

    def decorator(func):
        EXTRACTORS[extension] = func
        return func

    return decorator


@register_extractor(".pdf")
def extract_text_from_pdf(pdf_path, pages=None):
    """
    Extract text from a PDF file using pdfplumber.
    """
    text = []
    with pdfplumber.open(Path(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            if pages is None or i in pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
    return "\n".join(text)


@register_extractor(".txt")
def extract_text_from_txt(txt_path):
    """
    Extract text from a plain text file with encoding fallback.
    """
    try:
        with open(txt_path, encoding="utf-8") as file:
            return file.read()
    except UnicodeDecodeError:
        with open(txt_path, encoding="latin-1") as file:
            return file.read()


@register_extractor(".xlsx")
@register_extractor(".xls")
def extract_text_from_excel(excel_path):
    """
    Extract text from Excel workbooks (.xls/.xlsx).
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text.append(f"Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                text.append(row_text)
        text.append("")
    return "\n".join(text)


@register_extractor(".docx")
def extract_text_from_docx(docx_path):
    """
    Extract text from Microsoft Word documents (.docx).
    """
    doc = Document(docx_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


@register_extractor(".csv")
def extract_text_from_csv(csv_path, delimiter=","):
    """
    Extract text from CSV files.
    """
    text = []
    with open(csv_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile, delimiter=delimiter)
        for row in reader:
            row_text = " ".join(field for field in row if field.strip())
            if row_text:
                text.append(row_text)
    return "\n".join(text)


def extract_text(file_path, **kwargs):
    """
    Extract text from a file based on its extension.

    Args:
        file_path (str): Path to the file.
        **kwargs: Additional arguments for specific extractors.

    Returns:
        str: Extracted text.
    """
    ext = Path(file_path).suffix.lower()
    extractor = EXTRACTORS.get(ext)
    if not extractor:
        raise ValueError(f"No extractor registered for extension: {ext}")
    return extractor(file_path, **kwargs)


@tool
def extract_text_tool(input_args: ExtractTextInputs) -> str:
    """Create the extract_text tool."""
    return extract_text(
        input_args.file_path, pages=input_args.pages, delimiter=input_args.delimiter
    )
