"""Give It A Summary - Extraction Tools Module."""

import os
import tempfile
from pathlib import Path
from typing import Any

import pdfplumber
from docx import Document
from langchain_core.tools import tool
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.models.agent import ExtractTextInputs
from app.utilities.logs import get_logger

logger = get_logger(__name__)


class TextExtractionResult(BaseModel):
    """Result of text extraction operation."""

    content: str = Field(..., description="Extracted text content")
    metadata: dict[str, Any] = Field(default_factory=dict, description="Extraction metadata")
    success: bool = Field(default=True, description="Whether extraction was successful")
    error_message: str | None = Field(None, description="Error message if extraction failed")


def validate_file(file_path: str) -> bool:
    """Validate that file exists and is readable."""
    if not os.path.exists(file_path):
        logger.error("File does not exist: %s", file_path)
        return False

    if not os.path.isfile(file_path):
        logger.error("Path is not a file: %s", file_path)
        return False

    if not os.access(file_path, os.R_OK):
        logger.error("File is not readable: %s", file_path)
        return False

    return True


def get_file_size_mb(file_path: str) -> float:
    """Get file size in MB."""
    return os.path.getsize(file_path) / (1024 * 1024)


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF files using pdfplumber."""
    try:
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    except ImportError:
        logger.warning("pdfplumber not available, cannot extract PDF text")
        return ""
    except Exception as e:  # noqa: BLE001
        logger.error("Error extracting PDF text: %s", str(e))
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX files."""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except ImportError:
        logger.warning("python-docx not available, cannot extract DOCX text")
        return ""
    except Exception as e:  # noqa: BLE001
        logger.error("Error extracting DOCX text: %s", str(e))
        return ""


def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel files using openpyxl."""
    try:
        wb = load_workbook(file_path, read_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except ImportError:
        logger.warning("openpyxl not available, cannot extract Excel text")
        return ""
    except Exception as e:  # noqa: BLE001
        logger.error("Error extracting Excel text: %s", str(e))
        return ""


def extract_text_from_csv(file_path: str) -> str:
    """Extract text from CSV files."""
    try:
        import csv

        text = ""
        with open(file_path, encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                row_text = ",".join(str(cell) for cell in row)
                text += row_text + "\n"
        return text.strip()
    except Exception as e:  # noqa: BLE001
        logger.error("Error extracting CSV text: %s", str(e))
        return ""


def extract_text_from_txt(file_path: str) -> str:
    """Extract text from plain text files."""
    try:
        with open(file_path, encoding="utf-8", errors="ignore") as f:
            return f.read().strip()
    except Exception as e:  # noqa: BLE001
        logger.error("Error extracting TXT text: %s", str(e))
        return ""


def extract_text(file_path: str, **kwargs) -> str:
    """
    Extract text from various file formats.

    Supported formats: PDF, DOCX, TXT, XLSX, XLS, CSV

    Args:
        file_path: Path to the file to extract text from
        **kwargs: Additional arguments (ignored for compatibility)

    Returns:
        Extracted text content as string
    """
    file_ext = Path(file_path).suffix.lower()

    if file_ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif file_ext in [".docx", ".doc"]:
        return extract_text_from_docx(file_path)
    elif file_ext in [".xlsx", ".xls"]:
        return extract_text_from_excel(file_path)
    elif file_ext == ".csv":
        return extract_text_from_csv(file_path)
    elif file_ext == ".txt":
        return extract_text_from_txt(file_path)
    else:
        logger.warning("Unsupported file format: %s", file_ext)
        return extract_text_from_txt(file_path)


def extract_text_from_file(file_path: str, **kwargs) -> TextExtractionResult:
    """
    Extract text from various file formats using custom extraction functions.

    Supported formats: PDF, DOCX, TXT, XLSX, XLS, CSV

    Args:
        file_path: Path to the file to extract text from
        **kwargs: Additional arguments (ignored for compatibility)

    Returns:
        TextExtractionResult with extracted content and metadata
    """
    try:
        if not validate_file(file_path):
            return TextExtractionResult(
                content="", success=False, error_message="File validation failed"
            )

        file_size_mb = get_file_size_mb(file_path)
        if file_size_mb > 50:
            logger.warning("Large file detected: %.2f MB", file_size_mb)

        file_ext = Path(file_path).suffix.lower()

        logger.info("Extracting text from %s (format: %s)", file_path, file_ext)
        extracted_content = extract_text(file_path, **kwargs)

        if not extracted_content or not extracted_content.strip():
            logger.warning("No text content extracted from %s", file_path)
            return TextExtractionResult(
                content="",
                metadata={
                    "file_path": file_path,
                    "file_size_mb": file_size_mb,
                    "file_type": file_ext,
                },
                success=False,
                error_message="No text content found in file",
            )

        cleaned_content = " ".join(extracted_content.split())

        metadata = {
            "file_path": file_path,
            "file_size_mb": file_size_mb,
            "file_type": file_ext,
            "content_length": len(cleaned_content),
            "extraction_method": "custom_extractors",
        }

        logger.info("Successfully extracted %d characters from %s", len(cleaned_content), file_path)

        return TextExtractionResult(content=cleaned_content, metadata=metadata, success=True)

    except OSError as e:
        error_msg = f"Text extraction failed: {e}"
        logger.error("%s for file: %s", error_msg, file_path)
        return TextExtractionResult(
            content="",
            metadata={"file_path": file_path, "error_type": type(e).__name__},
            success=False,
            error_message=error_msg,
        )


def extract_text_from_bytes(file_bytes: bytes, filename: str, **kwargs) -> TextExtractionResult:
    """
    Extract text from file bytes (useful for uploaded files).

    Args:
        file_bytes: Raw file bytes
        filename: Original filename (used for format detection)
        **kwargs: Additional arguments (ignored for compatibility)

    Returns:
        TextExtractionResult with extracted content and metadata
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix) as temp_file:
            temp_file.write(file_bytes)
            temp_file_path = temp_file.name

        try:
            result = extract_text_from_file(temp_file_path, **kwargs)

            result.metadata["original_filename"] = filename
            result.metadata["temp_file_used"] = True

            return result

        finally:
            try:
                os.unlink(temp_file_path)
            except OSError as e:
                logger.warning("Failed to clean up temporary file %s: %s", temp_file_path, str(e))

    except OSError as e:
        error_msg = f"Text extraction from bytes failed: {e}"
        logger.error(error_msg)
        return TextExtractionResult(
            content="",
            metadata={"original_filename": filename, "error_type": type(e).__name__},
            success=False,
            error_message=error_msg,
        )


@tool
def extract_text_tool(inputs: ExtractTextInputs) -> str:
    """
    LangGraph tool for text extraction.

    This tool extracts text from various file formats and returns the content
    as a string for further processing in the agent workflow.

    Args:
        inputs: ExtractTextInputs containing file path and extraction parameters

    Returns:
        Extracted text content as string, or error message if extraction fails
    """
    logger.info("Extract text tool called with inputs: %s", inputs)

    try:
        result = extract_text_from_file(
            file_path=inputs.file_path,
        )

        if result.success:
            logger.info("Text extraction successful: %d characters extracted", len(result.content))
            return result.content
        else:
            error_msg = f"Text extraction failed: {result.error_message}"
            logger.error(error_msg)
            return error_msg

    except OSError as e:
        error_msg = f"Tool execution failed: {e}"
        logger.error(error_msg)
        return error_msg
