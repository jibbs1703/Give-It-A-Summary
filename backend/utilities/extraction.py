"""Give It A Summary - Extraction Utilities Module."""

import pdfplumber
from openpyxl import load_workbook


def extract_text_from_pdf(pdf_path):
    """
    Extract text from a PDF file.
    """
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_text_from_txt(txt_path):
    """
    Extract text from a .txt file.
    """
    with open(txt_path, encoding='utf-8') as file:
        text = file.read()
    return text


def extract_text_from_excel(excel_path):
    """
    Extract text from an excel file.
    """
    wb = load_workbook(excel_path)
    text = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text += f"Sheet: {sheet_name}\n"
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                text += row_text + "\n"
        text += "\n"
    return text
